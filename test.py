import openpyxl
import os
import shutil

'''
workbook = openpyxl.load_workbook('3차년도 호흡음체크.xlsx')
worksheet = workbook['호흡음']

cell = worksheet['B2'].value
# cell = worksheet.cell(row=3, column=3).value
print(cell[23])
print(type(cell))

cell_range = worksheet['C2':'C5']
print(cell_range)

value_lst = []
for row in cell_range:
    for cell in row:
        value_lst.append(cell.value)

print(value_lst)

'''
workbook = openpyxl.load_workbook('3차년도 호흡음체크.xlsx')
worksheet = workbook['호흡음']

window = '0.4_0.2'

for cell, start, end in [[2, 2, 5], [3, 2, 3], [8, 8, 8], [9, 8, 11], [17, 17, 20], [18, 17, 22], [24, 24, 27],
                         [25, 24, 27],
                         [29, 29, 30], [30, 29, 34], [37, 37, 42], [38, 37, 42], [43, 43, 45], [44, 43, 46],
                         [62, 62, 63], [63, 62, 63],
                         [65, 65, 68], [66, 65, 70], [74, 74, 76], [75, 74, 77], [86, 86, 89], [87, 86, 88],
                         [95, 95, 97],
                         [96, 95, 96], [107, 106, 110], [122, 122, 127], [123, 122, 124], [128, 128, 131],
                         [129, 128, 130],
                         [134, 134, 137], [135, 134, 136], [140, 140, 145], [141, 140, 144], [147, 147, 151],
                         [148, 147, 152],
                         [153, 153, 156], [154, 153, 155], [159, 159, 163], [149, 150, 153], [160, 159, 160],
                         [164, 164, 166],
                         [165, 164, 166], [174, 174, 179], [175, 174, 178], [180, 180, 185], [181, 180, 185],
                         [188, 188, 191],
                         [189, 188, 191], [193, 193, 198], [194, 193, 197], [199, 199, 204], [200, 199, 204],
                         [205, 205, 207],
                         [206, 205, 207], [212, 212, 216], [213, 212, 216], [218, 218, 221], [219, 218, 219],
                         [224, 224, 225],
                         [225, 224, 229], [230, 230, 235], [231, 230, 236], [245, 245, 251], [246, 245, 250],
                         [253, 253, 255],
                         [254, 253, 253], [257, 257, 258], [258, 257, 258], [267, 267, 270], [268, 268, 272],
                         [279, 279, 281],
                         [280, 279, 281], [282, 282, 286], [283, 282, 288], [289, 289, 294], [295, 295, 299],
                         [296, 295, 299],
                         [300, 300, 306], [301, 300, 307], [308, 308, 311], [309, 308, 310], [312, 312, 313],
                         [315, 314, 314],
                         [320, 320, 322], [321, 320, 322], [323, 323, 323], [325, 325, 327], [326, 325, 331],
                         [332, 332, 335],
                         [333, 323, 324], [336, 336, 338], [337, 336, 341], [342, 342, 344], [343, 342, 343],
                         [345, 345, 350],
                         [346, 345, 348], [352, 351, 353], [363, 363, 365]]:

    file_name = worksheet['B' + str(cell)].value
    cell_start = start
    cell_end = end

    f_code = file_name[23]

    start_cell_range = []
    end_cell_range = []

    if f_code == 'C':
        start_cell_range = worksheet['C' + str(cell_start):'C' + str(cell_end)]
        end_cell_range = worksheet['D' + str(cell_start):'D' + str(cell_end)]
    elif f_code == 'D':
        start_cell_range = worksheet['F' + str(cell_start):'F' + str(cell_end)]
        end_cell_range = worksheet['G' + str(cell_start):'G' + str(cell_end)]

    start_time_lst = []
    end_time_lst = []

    for row in start_cell_range:
        for cell in row:
            start_time_lst.append(cell.value)

    for row in end_cell_range:
        for cell in row:
            end_time_lst.append(cell.value)

    memo_lst = []
    memo_range = []
    if f_code == 'C':
        memo_range = worksheet['E' + str(cell_start):'E' + str(cell_end)]
    elif f_code == 'D':
        memo_range = worksheet['H' + str(cell_start):'H' + str(cell_end)]

    for row in memo_range:
        for cell in row:
            memo_lst.append(cell.value)

    wheeze_file_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/wheeze'
    crackle_file_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/crackle'
    normal_file_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/normal'

    wheeze_save_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/분류/test/wheeze'
    crackle_save_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/분류/test/crackle'
    normal_save_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/분류/test/normal'

    # ************************
    file_number = file_name[4:22]
    # ************************

    for path in [wheeze_file_path, crackle_file_path]:
            if 'wheeze' in file_name:
                file_list = os.listdir(path)
                print('file_list : ' + str(file_list))
                for f_name in file_list:
                    if file_number in f_name:
                        f_name_lst = f_name.split('_')
                        f_start_time = float(f_name_lst[-2])
                        f_end_time = float(f_name_lst[-1][:-4])

                        print('file time range : ' + str(f_start_time) + ' ~ ' + str(f_end_time))

                        save_flag = 0

                        for idx in range(len(start_cell_range)):
                            start_time = start_time_lst[idx]
                            end_time = end_time_lst[idx]

                            print('excel time range : ' + str(start_time) + ' ~ ' + str(end_time))

                            if f_start_time <= start_time:
                                if start_time < f_end_time <= end_time:
                                    print(f_name + ' - wheeze')
                                    shutil.copy(path + '/' + f_name, wheeze_save_path)
                                    save_flag = 1
                                elif end_time < f_end_time:
                                    print(f_name + ' - wheeze')
                                    shutil.copy(path + '/' + f_name, wheeze_save_path)
                                    save_flag = 1
                            elif start_time <= f_start_time < end_time:
                                print(f_name + ' - wheeze')
                                shutil.copy(path + '/' + f_name, wheeze_save_path)
                                save_flag = 1

                        if save_flag == 0:
                            print(f_name + ' - normal')
                            shutil.copy(path + '/' + f_name, normal_save_path)

            elif 'crackle' in file_name:
                file_list = os.listdir(path)

                for f_name in file_list:
                    if file_number in f_name:
                        f_name_lst = f_name.split('_')
                        f_start_time = float(f_name_lst[-2])
                        f_end_time = float(f_name_lst[-1][:-4])

                        print('file time range : ' + str(f_start_time) + ' ~ ' + str(f_end_time))

                        save_flag = 0

                        for idx in range(len(start_cell_range)):
                            start_time = start_time_lst[idx]
                            end_time = end_time_lst[idx]

                            print('excel time range : ' + str(start_time) + ' ~ ' + str(end_time))

                            if f_start_time <= start_time:
                                if start_time < f_end_time <= end_time:
                                    print(f_name + ' - crackle')
                                    shutil.copy(path + '/' + f_name, wheeze_save_path)
                                    save_flag = 1
                                elif end_time < f_end_time:
                                    print(f_name + ' - crackle')
                                    shutil.copy(path + '/' + f_name, wheeze_save_path)
                                    save_flag = 1
                            elif start_time <= f_start_time < end_time:
                                print(f_name + ' - crackle')
                                shutil.copy(path + '/' + f_name, wheeze_save_path)
                                save_flag = 1

                        if save_flag == 0:
                            print(f_name + ' - normal')
                            shutil.copy(path + '/' + f_name, normal_save_path)