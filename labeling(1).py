import openpyxl
import os
import shutil

'''
workbook = openpyxl.load_workbook('3차년도 호흡음체크.xlsx')
worksheet = workbook['호흡음']

cell = worksheet['C2'].value
# cell = worksheet.cell(row=3, column=3).value
print(cell)

print(type(cell))

cell_range = worksheet['C2':'C5']

value_lst = []
for row in cell_range:
    for cell in row:
        value_lst.append(cell.value)

print(value_lst)

'''
workbook = openpyxl.load_workbook('190523-호흡음체크-확인답변.xlsx')
worksheet = workbook['호흡음']

window = '2.0_1.0'

for cell, start, end in [[2, 3, 6], [8, 9, 15], [9, 9, 16], [17, 18, 21], [18, 18, 23], [24, 25, 28], [25, 25, 28],
                         [29, 30, 36], [30, 30, 36], [37, 38, 41], [42, 43, 49], [43, 43, 48], [50, 51, 59],
                         [51, 51, 60], [61, 62, 62], [62, 62, 63], [64, 65, 72], [65, 65, 70], [73, 74, 78],
                         [74, 74, 78], [79, 80, 84], [80, 80, 83], [85, 86, 93], [86, 86, 91], [94, 95, 100],
                         [95, 95, 104], [106, 106, 107], [108, 109, 118], [109, 109, 115], [120, 120, 121],
                         [122, 123, 127], [123, 123, 127], [128, 129, 133], [129, 129, 133], [135, 135, 138],
                         [139, 140, 142], [143, 144, 147], [144, 144, 148], [149, 150, 153], [150, 150, 152],
                         [154, 155, 163], [155, 155, 158], [165, 165, 169], [170, 171, 177], [171, 171, 177],
                         [178, 179, 181], [179, 179, 182], [183, 184, 188], [184, 184, 187], [189, 190, 193],
                         [194, 195, 196], [195, 195, 200], [201, 202, 213], [202, 202, 209], [214, 215, 217],
                         [220, 221, 225], [221, 221, 225], [226, 227, 230], [227, 227, 234], [237, 238, 241],
                         [238, 238, 241], [242, 243, 245], [243, 243, 245], [246, 247, 255], [247, 247, 250],
                         [256, 257, 267], [257, 257, 262]]:

    file_name = worksheet['B' + str(cell)].value
    print(file_name)
    cell_start = start
    cell_end = end

    f_code = file_name[20]
    start_cell_range = []
    end_cell_range = []

    if f_code == 'C':
        start_cell_range = worksheet['C' + str(cell_start):'C' + str(cell_end)]
        end_cell_range = worksheet['D' + str(cell_start):'D' + str(cell_end)]
    elif f_code == 'D':
        start_cell_range = worksheet['F' + str(cell_start):'F' + str(cell_end)]
        end_cell_range = worksheet['G' + str(cell_start):'G' + str(cell_end)]

    start_time_lst = []
    end_time_lst = []

    # ************************
    for row in start_cell_range:
        for cell in row:
            start_time_lst.append(cell.value)

    for row in end_cell_range:
        for cell in row:
            end_time_lst.append(cell.value)
    # ************************

    memo_lst = []
    memo_range = []
    if f_code == 'C':
        memo_range = worksheet['E' + str(cell_start):'E' + str(cell_end)]
    elif f_code == 'D':
        memo_range = worksheet['H' + str(cell_start):'H' + str(cell_end)]

    for row in memo_range:
        for cell in row:
            memo_lst.append(cell.value)

    wheeze_file_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/wheeze'
    crackle_file_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/crackle'
    normal_file_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/normal'

    wheeze_save_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/분류/wheeze'
    crackle_save_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/분류/crackle'
    normal_save_path = 'E:/experiment/2020.09_LungSound/1.dataset/' + window + '/분류/normal'

    file_number = file_name[4:21]

    for path in [wheeze_file_path, crackle_file_path]:
        if ('wheeze' in file_name) and ('crackle' in file_name):
            file_list = os.listdir(path)
            for f_name in file_list:
                if file_number in f_name:
                    #shutil.copy(path + '/' + f_name, 'C:/Users/SAMSUNG/바탕 화면/2.0_1.0/2.0_1.0/test_wheeze_crackle')
                    print('Save wheeze, crackle')
                    f_name_lst = f_name.split('_')
                    f_start_time = float(f_name_lst[-2])
                    f_end_time = float(f_name_lst[-1][:-4])

                    # print('file time range : ' + str(f_start_time) + ' ~ ' + str(f_end_time))

                    save_flag = 0

                    for idx in range(len(start_cell_range)):
                        start_time = start_time_lst[idx].second + start_time_lst[idx].microsecond * 0.000001
                        end_time = end_time_lst[idx].second + end_time_lst[idx].microsecond * 0.000001

                        # print('excel time range : ' + str(start_time) + ' ~ ' + str(end_time))

                        if f_start_time <= start_time:
                            if start_time < f_end_time <= end_time:
                                if memo_lst[idx] == 'wheeze':
                                    # print(f_name + ' - wheeze')
                                    shutil.copy(path + '/' + f_name, wheeze_save_path)
                                elif memo_lst[idx] == 'crackle':
                                    # print(f_name + ' - crackle')
                                    shutil.copy(path + '/' + f_name, crackle_save_path)
                                save_flag = 1
                            elif end_time < f_end_time:
                                if memo_lst[idx] == 'wheeze':
                                    # print(f_name + ' - wheeze')
                                    shutil.copy(path + '/' + f_name, wheeze_save_path)
                                elif memo_lst[idx] == 'crackle':
                                    # print(f_name + ' - crackle')
                                    shutil.copy(path + '/' + f_name, crackle_save_path)
                                save_flag = 1
                        elif start_time <= f_start_time < end_time:
                            if memo_lst[idx] == 'wheeze':
                                # print(f_name + ' - wheeze')
                                shutil.copy(path + '/' + f_name, wheeze_save_path)
                            elif memo_lst[idx] == 'crackle':
                                # print(f_name + ' - crackle')
                                shutil.copy(path + '/' + f_name, crackle_save_path)
                            save_flag = 1

                    if save_flag == 0:
                        # print(f_name + ' - normal')
                        shutil.copy(path + '/' + f_name, normal_save_path)

        else:
            if 'wheeze' in file_name:
                file_list = os.listdir(path)

                for f_name in file_list:
                    if file_number in f_name:
                        #shutil.copy(path + '/' + f_name, 'C:/Users/minwo/바탕 화면/2.0_1.0/2.0_1.0/test_wheeze')
                        print('Save wheeze')
                        f_name_lst = f_name.split('_')
                        f_start_time = float(f_name_lst[-2])
                        f_end_time = float(f_name_lst[-1][:-4])

                        # print('file time range : ' + str(f_start_time) + ' ~ ' + str(f_end_time))

                        save_flag = 0

                        for idx in range(len(start_cell_range)):
                            start_time = start_time_lst[idx].second + start_time_lst[idx].microsecond * 0.000001
                            end_time = end_time_lst[idx].second + end_time_lst[idx].microsecond * 0.000001

                            # print('excel time range : ' + str(start_time) + ' ~ ' + str(end_time))

                            if f_start_time <= start_time:
                                if start_time < f_end_time <= end_time:
                                    # print(f_name + ' - wheeze')
                                    shutil.copy(path + '/' + f_name, wheeze_save_path)
                                    save_flag = 1
                                elif end_time < f_end_time:
                                    # print(f_name + ' - wheeze')
                                    shutil.copy(path + '/' + f_name, wheeze_save_path)
                                    save_flag = 1
                            elif start_time <= f_start_time < end_time:
                                # print(f_name + ' - wheeze')
                                shutil.copy(path + '/' + f_name, wheeze_save_path)
                                save_flag = 1

                        if save_flag == 0:
                            # print(f_name + ' - normal')
                            shutil.copy(path + '/' + f_name, normal_save_path)

            elif 'crackle' in file_name:
                file_list = os.listdir(path)

                for f_name in file_list:
                    if file_number in f_name:
                        #shutil.copy(path + '/' + f_name, 'C:/Users/minwo/바탕 화면/2.0_1.0/2.0_1.0/test_crackle')
                        print('Save crackle')
                        f_name_lst = f_name.split('_')
                        f_start_time = float(f_name_lst[-2])
                        f_end_time = float(f_name_lst[-1][:-4])

                        # print('file time range : ' + str(f_start_time) + ' ~ ' + str(f_end_time))

                        save_flag = 0

                        for idx in range(len(start_cell_range)):
                            start_time = start_time_lst[idx].second + start_time_lst[idx].microsecond * 0.000001
                            end_time = end_time_lst[idx].second + end_time_lst[idx].microsecond * 0.000001

                            # print('excel time range : ' + str(start_time) + ' ~ ' + str(end_time))

                            if f_start_time <= start_time:
                                if start_time < f_end_time <= end_time:
                                    # print(f_name + ' - crackle')
                                    shutil.copy(path + '/' + f_name, crackle_save_path)
                                    save_flag = 1
                                elif end_time < f_end_time:
                                    # print(f_name + ' - crackle')
                                    shutil.copy(path + '/' + f_name, crackle_save_path)
                                    save_flag = 1
                            elif start_time <= f_start_time < end_time:
                                # print(f_name + ' - crackle')
                                shutil.copy(path + '/' + f_name, crackle_save_path)
                                save_flag = 1

                        if save_flag == 0:
                            # print(f_name + ' - normal')
                            shutil.copy(path + '/' + f_name, normal_save_path)
